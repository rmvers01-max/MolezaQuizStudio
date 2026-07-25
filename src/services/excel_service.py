from openpyxl import load_workbook
from src.models.pergunta import Pergunta


class ExcelService:

    def carregar(self, caminho):

        workbook = load_workbook(caminho)

        planilha = workbook.active

        perguntas = []

        # Começa na linha 2 porque a linha 1 é o cabeçalho
        for linha in planilha.iter_rows(min_row=2, values_only=True):

            pergunta = Pergunta(
                texto=linha[0],
                opcao_a=linha[1],
                opcao_b=linha[2],
                resposta=linha[3],
                imagem_a=linha[4],
                imagem_b=linha[5]
            )

            perguntas.append(pergunta)

        return perguntas
